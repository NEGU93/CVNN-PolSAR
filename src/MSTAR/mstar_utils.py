import numpy as np
from pathlib import Path
import matplotlib.pyplot as plt
from pdb import set_trace
import os
from sklearn import preprocessing
import matplotlib.image as mpimg
from openpyxl import load_workbook, Workbook
from openpyxl.worksheet.table import Table
from typing import List, Optional, Union

t_path = Union[str, Path]

root_path = Path('/media/barrachina/data/datasets/MSTAR/')


def plot_data(data):
    data = np.abs(data)
    arr = np.asarray(data)
    plt.imshow(arr, cmap='gray')
    plt.show()


def plot_jpg(mstar_data):
    fullfilename = str(root_path) + mstar_data[0]['path'] + '/' + mstar_data[0]['Filename'].upper()
    jpg_path = os.path.splitext(str(fullfilename))[0] + '.JPG'
    img = mpimg.imread(jpg_path)
    imgplot = plt.imshow(img, cmap='gray')
    plt.show()


def plot_history(history):
    plt.plot(history.history['accuracy'])
    plt.plot(history.history['val_accuracy'])
    plt.title('model accuracy')
    plt.ylabel('accuracy')
    plt.xlabel('epoch')
    plt.legend(['train', 'validation'], loc='upper left')
    plt.show()


def create_excel_file(fieldnames: List[str], row_data: List, filename: Optional[t_path] = None,
                      percentage_cols: Optional[List[str]] = None):
    if filename is None:
        filename = './log/runs_summary.xlsx'
    file_exists = os.path.isfile(filename)
    if file_exists:
        wb = load_workbook(filename)
        ws = wb.worksheets[0]
        del ws.tables["Table1"]
    else:
        wb = Workbook()
        ws = wb.worksheets[0]
        ws.append(fieldnames)
    ws.append(row_data)
    # TODO: What if len(row_data) is longer than the dictionary? It corresponds with excel's column names?
    tab = Table(displayName="Table1", ref="A1:" + str(chr(64 + len(row_data))) + str(ws.max_row))
    if percentage_cols is not None:
        for col in percentage_cols:
            ws[col + str(ws.max_row)].number_format = '0.00%'
    ws.add_table(tab)
    wb.save(filename)


if __name__ == '__main__':
    mstar_data = np.load(root_path/'data.npy', allow_pickle=True)
    plot_data(mstar_data[0]['data'])
    plot_jpg(mstar_data)


